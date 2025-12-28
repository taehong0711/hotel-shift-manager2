import streamlit as st
import pandas as pd
from ortools.sat.python import cp_model
import io
import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# =========================================================
# ✅ 0) VERSION (key busting)
# =========================================================
APP_VERSION = "2026.01.28.v3"   # <- 형이 수정할 때마다 문자열 바꾸면 100% 새로 반영됨

def versioned(key: str) -> str:
    return f"{key}__{APP_VERSION}"

def reset_all_except_password():
    keep = {}
    if st.session_state.get("password_correct") is True:
        keep["password_correct"] = True
    st.cache_data.clear()
    st.session_state.clear()
    for k, v in keep.items():
        st.session_state[k] = v
    st.session_state["app_version"] = APP_VERSION
    st.rerun()

def ensure_version_fresh():
    # 코드 교체(=APP_VERSION 변경)되면 자동으로 세션/캐시 클리어
    if st.session_state.get("app_version") != APP_VERSION:
        keep_pw = st.session_state.get("password_correct", False)
        st.cache_data.clear()
        st.session_state.clear()
        if keep_pw:
            st.session_state["password_correct"] = True
        st.session_state["app_version"] = APP_VERSION

# =========================================================
# 0) Page + Password
# =========================================================
st.set_page_config(layout="wide", page_title="ホテルシフト自動作成 Pro (2026 Ver / 2-Stage)")

SECRET_PASSWORD = st.secrets["password"] if "password" in st.secrets else "1234"

def check_password():
    def password_entered():
        if st.session_state[versioned("password")] == SECRET_PASSWORD:
            st.session_state["password_correct"] = True
            del st.session_state[versioned("password")]
        else:
            st.session_state["password_correct"] = False

    if "password_correct" not in st.session_state:
        st.text_input("パスワードを入力してください (Password)", type="password",
                      on_change=password_entered, key=versioned("password"))
        return False
    elif not st.session_state["password_correct"]:
        st.text_input("パスワードを入力してください (Password)", type="password",
                      on_change=password_entered, key=versioned("password"))
        st.error("パスワードが間違っています。")
        return False
    else:
        return True

if not check_password():
    st.stop()

# ✅ password 통과 후 버전 체크/초기화
ensure_version_fresh()

# =========================================================
# 1) Base Config
# =========================================================
OFF_CODE = "公"
MYONG_CODE = "-"     # 明け
UNASSIGNED_CODE = "未"  # Stage1 내부용(표시는 빈칸)
WEEKDAY_CHARS = ["月", "火", "水", "木", "金", "土", "日"]

def remove_D_from_shift_lists():
    st.session_state["shifts_day"] = [c for c in st.session_state["shifts_day"] if c != "D"]
    st.session_state["shifts_night"] = [c for c in st.session_state["shifts_night"] if c != "D"]

# 최초 진입(또는 강제리셋) 시에만 기본값 세팅
if "init_done" not in st.session_state:
    st.session_state["shifts_day"] = ["E1", "E2", "G1", "G1U", "H1", "H2", "I1", "I2", "L1"]
    st.session_state["shifts_night"] = ["Q1", "X1", "R1"]  # ✅ 야근 3코드 (각 1명/일)
    st.session_state["init_done"] = True

remove_D_from_shift_lists()

SPECIAL_CODES = ["日", MYONG_CODE, OFF_CODE]
SPECIAL_CODES_STAGE1 = ["日", MYONG_CODE, OFF_CODE, UNASSIGNED_CODE]

# =========================================================
# Staff DB (야근 가능코드 반영)
# =========================================================
INITIAL_STAFF_DB = [
    {"name": "井戸",   "gender": "M", "role": "Manager", "target_off": 8, "skills": "日, G1, H1, X1, -, 公"},
    {"name": "畑瀬",   "gender": "M", "role": "Manager", "target_off": 8, "skills": "日, G1, H1, X1, Q1, -, 公"},
    {"name": "夏川",   "gender": "F", "role": "Manager", "target_off": 8, "skills": "E1, 公"},
    {"name": "都筑",   "gender": "M", "role": "Manager", "target_off": 8, "skills": "日, G1, H1, X1, -, 公"},
    {"name": "山口",   "gender": "M", "role": "Manager", "target_off": 8, "skills": "日, G1, H1, X1, -, 公"},

    {"name": "茅島",   "gender": "F", "role": "Staff",   "target_off": 8, "skills": "日, G1U, H1, I1, I2, X1, Q1, -, 公"},
    {"name": "馬場",   "gender": "F", "role": "Staff",   "target_off": 8, "skills": "日, G1U, H1, I1, I2, Q1, -, 公"},
    {"name": "池田",   "gender": "F", "role": "Staff",   "target_off": 8, "skills": "日, G1U, H1, H2, I1, I2, L1, Q1, R1, -, 公"},
    {"name": "川野",   "gender": "F", "role": "Staff",   "target_off": 8, "skills": "日, G1U, H1, H2, I1, I2, Q1, R1, -, 公"},
    {"name": "加藤",   "gender": "F", "role": "Staff",   "target_off": 8, "skills": "日, G1U, H1, H2, I1, I2, L1, -, 公"},
    {"name": "四ヶ所", "gender": "F", "role": "Staff",   "target_off": 8, "skills": "日, G1U, H1, H2, I1, I2, L1, Q1, R1, -, 公"},
    {"name": "朴",     "gender": "M", "role": "Staff",   "target_off": 8, "skills": "日, G1U, H1, H2, I1, I2, L1, X1, R1, -, 公"},
    {"name": "ハマノ", "gender": "F", "role": "Staff",   "target_off": 8, "skills": "日, G1U, H1, H2, I1, I2, L1, 公"},
    {"name": "田中",   "gender": "M", "role": "Staff",   "target_off": 8, "skills": "日, G1U, H1, H2, I1, I2, L1, R1, -, 公"},
    {"name": "市之瀬", "gender": "F", "role": "Staff",   "target_off": 8, "skills": "日, G1U, H1, H2, I1, I2, L1, R1, -, 公"},
    {"name": "鬼塚",   "gender": "F", "role": "Staff",   "target_off": 8, "skills": "日, G1U, H1, H2, I1, I2, 公"},
    {"name": "春山",   "gender": "F", "role": "Staff",   "target_off": 8, "skills": "日, G1U, H1, H2, I1, I2, L1, 公"},
    {"name": "佐伯",   "gender": "F", "role": "Staff",   "target_off": 8, "skills": "E2, 公"},
    {"name": "杉浦",   "gender": "F", "role": "Staff",   "target_off": 8, "skills": "日, G1U, H1, H2, I1, I2, L1, 公"},
    {"name": "野田",   "gender": "F", "role": "Staff",   "target_off": 8, "skills": "E1, 公"},
]

# =========================================================
# Helpers
# =========================================================
def norm_code(x):
    if pd.isna(x):
        return ""
    s = str(x).strip()
    if s == "":
        return ""
    if s.upper() == "OFF" or s in ["休", "公休"]:
        return OFF_CODE
    if s == "明":
        return MYONG_CODE
    if s == "D":
        return ""   # ✅ D 제거
    return s

def build_day_headers(year, month, num_days):
    headers = []
    for d in range(num_days):
        cur_date = datetime.date(year, month, d + 1)
        w_str = WEEKDAY_CHARS[cur_date.weekday()]
        headers.append(f"{d + 1}日({w_str})")
    return headers

def parse_skills(skill_str: str):
    if skill_str is None:
        return set()
    s = str(skill_str).replace("明", MYONG_CODE).replace("OFF", OFF_CODE)
    items = [x.strip() for x in s.split(",") if x.strip()]
    items = [x for x in items if x != "D"]
    return set(items)

def validate_mandatory_coverage(staff_data, shifts_day, shifts_night):
    required = list(shifts_night) + (["L1"] if "L1" in shifts_day else [])
    skill_map = {s["name"]: parse_skills(s.get("skills", "")) for s in staff_data}
    missing = []
    for code in required:
        eligible = [name for name, sk in skill_map.items() if code in sk]
        if len(eligible) == 0:
            missing.append(code)
    return missing

def summarize_requests(requests, shifts_day, shifts_night):
    cnt_off = cnt_night = cnt_l1 = cnt_daywish = cnt_nichi = 0
    for _, mp in requests.items():
        for _, code in mp.items():
            if code == OFF_CODE:
                cnt_off += 1
            elif code in shifts_night:
                cnt_night += 1
            elif code == "L1":
                cnt_l1 += 1
            elif code == "日":
                cnt_nichi += 1
            elif code in shifts_day and code != "L1":
                cnt_daywish += 1
    return {
        "希望休(公)": cnt_off,
        "希望勤務(日勤)": cnt_daywish,
        "夜勤希望(Q1/X1/R1)": cnt_night,
        "L1希望": cnt_l1,
        "日希望": cnt_nichi,
    }

# =========================================================
# Excel Styling
# =========================================================
def create_styled_excel(df_shift, df_summary, requests, year, month):
    wb = Workbook()
    ws_shift = wb.active
    ws_shift.title = "Shift"

    for r in dataframe_to_rows(df_shift, index=False, header=True):
        ws_shift.append(r)

    fill_off = PatternFill(start_color="F0F2F6", end_color="F0F2F6", fill_type="solid")
    fill_night = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    fill_myong = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    fill_l1 = PatternFill(start_color="E1BEE7", end_color="E1BEE7", fill_type="solid")
    fill_nichi = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")

    fill_sat_bg = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    fill_sun_bg = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    center_align = Alignment(horizontal="center", vertical="center")

    font_req = Font(bold=True, color="0000FF")
    font_sat = Font(bold=True, color="0000FF")
    font_sun = Font(bold=True, color="FF0000")

    night_codes = st.session_state["shifts_night"]

    header_row = ws_shift[1]
    for cell in header_row:
        cell.alignment = center_align
        cell.border = thin_border
        val = str(cell.value)
        if "(" in val:
            if "(土)" in val:
                cell.font = font_sat
                cell.fill = fill_sat_bg
            elif "(日)" in val:
                cell.font = font_sun
                cell.fill = fill_sun_bg

    for row in ws_shift.iter_rows(min_row=2, max_row=ws_shift.max_row, min_col=1, max_col=ws_shift.max_column):
        staff_name = str(row[0].value)
        for cell in row:
            cell.alignment = center_align
            cell.border = thin_border
            val = str(cell.value)
            col_idx = cell.column

            if val == OFF_CODE:
                cell.fill = fill_off
                cell.font = Font(color="BDC3C7")
            elif val in night_codes:
                cell.fill = fill_night
                cell.font = Font(color="B71C1C")
            elif val == MYONG_CODE:
                cell.fill = fill_myong
                cell.font = Font(color="F57F17")
            elif val == "L1":
                cell.fill = fill_l1
            elif val == "日":
                cell.fill = fill_nichi
                cell.font = Font(bold=True)

            if col_idx > 2:
                day_num = col_idx - 2
                if staff_name in requests and day_num in requests[staff_name]:
                    if requests[staff_name][day_num] == val:
                        cell.font = font_req

    ws_summary = wb.create_sheet("Summary")
    for r in dataframe_to_rows(df_summary, index=False, header=True):
        ws_summary.append(r)

    fill_alert = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    for row in ws_summary.iter_rows(min_row=2, max_row=ws_summary.max_row, min_col=2, max_col=ws_summary.max_column):
        for cell in row:
            cell.alignment = center_align
            cell.border = thin_border
            if cell.value == 0:
                cell.fill = fill_alert
                cell.font = Font(color="FF0000", bold=True)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# =========================================================
# HTML Table
# =========================================================
def generate_colored_table_html(df, requests):
    html = '<div style="overflow-x: auto; font-family: sans-serif; font-size: 0.9em;">'
    html += '<table style="border-collapse: collapse; width: 100%; white-space: nowrap;">'

    html += '<thead><tr style="background-color: #f8f9fa;">'
    for col in df.columns:
        bg_style = ""
        text_color = "black"
        if "(土)" in col:
            bg_style = "background-color: #D6EAF8;"
            text_color = "blue"
        elif "(日)" in col:
            bg_style = "background-color: #FADBD8;"
            text_color = "red"
        html += f'<th style="border: 1px solid #ddd; padding: 8px; {bg_style} color: {text_color}; text-align: center; position: sticky; top: 0; z-index: 2;">{col}</th>'
    html += "</tr></thead>"

    html += "<tbody>"
    for _, row in df.iterrows():
        html += "<tr>"
        staff_name = row["Staff"]
        for col_name, val in row.items():
            bg_color = "white"
            color = "black"
            font_weight = "normal"
            border_style = "1px solid #ddd"

            if val == OFF_CODE:
                bg_color = "#f0f2f6"
                color = "#bdc3c7"
            elif val in st.session_state["shifts_night"]:
                bg_color = "#ffcdd2"
                color = "#b71c1c"
            elif val == MYONG_CODE:
                bg_color = "#fff9c4"
                color = "#f57f17"
            elif val == "L1":
                bg_color = "#e1bee7"
            elif val == "日":
                bg_color = "#c8e6c9"
                font_weight = "bold"

            if "日(" in col_name:
                day_str = col_name.split("日")[0]
                if day_str.isdigit():
                    day_num = int(day_str)
                    if staff_name in requests and day_num in requests[staff_name]:
                        if requests[staff_name][day_num] == val:
                            border_style = "2px solid blue"
                            font_weight = "bold"

            html += f'<td style="border: {border_style}; padding: 6px; background-color: {bg_color}; color: {color}; font-weight: {font_weight}; text-align: center;">{val}</td>'
        html += "</tr>"
    html += "</tbody></table></div>"
    return html

# =========================================================
# Solver (2-Stage)
# =========================================================
@st.cache_data(show_spinner=False)
def solve_stage1(num_days, year, month, prev_history, requests, staff_data,
                shifts_day, shifts_night, closed_days, _version_stamp: str):
    """
    Stage1:
    - 입력된 (公/희망근무/야근/L1/日) 하드 고정
    - ✅ 야근(Q1,X1,R1) 매일 각각 1명 하드
    - ✅ L1 매일 1명 하드
    - 나머지 주간은 未로 남기고 표시상 빈칸
    """
    model = cp_model.CpModel()
    ALL_SHIFTS = shifts_day + shifts_night + SPECIAL_CODES_STAGE1
    staff_indices = range(len(staff_data))
    days_indices = range(num_days)
    closed_idx = set([d - 1 for d in closed_days if 1 <= d <= num_days])

    shifts = {}
    for s in staff_indices:
        for d in days_indices:
            for code in ALL_SHIFTS:
                shifts[(s, d, code)] = model.NewBoolVar(f"s{s}_d{d}_{code}")

    for s in staff_indices:
        for d in days_indices:
            model.Add(sum(shifts[(s, d, c)] for c in ALL_SHIFTS) == 1)

    for s in staff_indices:
        allowed = parse_skills(staff_data[s].get("skills", "")) | {UNASSIGNED_CODE}
        for d in days_indices:
            for code in ALL_SHIFTS:
                if code not in allowed:
                    model.Add(shifts[(s, d, code)] == 0)

    # prev month carry
    for s_idx, staff in enumerate(staff_data):
        name = staff["name"]
        h_d1 = norm_code(prev_history.get(name, {}).get("d-1", OFF_CODE))
        h_d2 = norm_code(prev_history.get(name, {}).get("d-2", OFF_CODE))
        h_d3 = norm_code(prev_history.get(name, {}).get("d-3", OFF_CODE))

        if h_d1 in shifts_night:
            model.Add(shifts[(s_idx, 0, MYONG_CODE)] == 1)

        if h_d1 == MYONG_CODE:
            for day_code in shifts_day + ["日", MYONG_CODE]:
                if day_code in ALL_SHIFTS:
                    model.Add(shifts[(s_idx, 0, day_code)] == 0)

        w_d3 = 1 if h_d3 != OFF_CODE else 0
        w_d2 = 1 if h_d2 != OFF_CODE else 0
        w_d1 = 1 if h_d1 != OFF_CODE else 0
        c0 = 1 - shifts[(s_idx, 0, OFF_CODE)] if 0 < num_days else 0
        c1 = 1 - shifts[(s_idx, 1, OFF_CODE)] if 1 < num_days else 0

        model.Add(w_d3 + w_d2 + w_d1 + c0 + c1 <= 4)
        if num_days >= 3:
            c2 = 1 - shifts[(s_idx, 2, OFF_CODE)]
            model.Add(w_d2 + w_d1 + c0 + c1 + c2 <= 4)
        if num_days >= 4:
            c2 = 1 - shifts[(s_idx, 2, OFF_CODE)]
            c3 = 1 - shifts[(s_idx, 3, OFF_CODE)]
            model.Add(w_d1 + c0 + c1 + c2 + c3 <= 4)

    # night -> next day is 明(-)
    for s in staff_indices:
        for d in range(num_days - 1):
            is_night = sum(shifts[(s, d, c)] for c in shifts_night if c in ALL_SHIFTS)
            model.Add(shifts[(s, d + 1, MYONG_CODE)] == is_night)

    # 明(-) -> next day cannot be day shift / 日 / 明
    for s in staff_indices:
        for d in range(num_days - 1):
            for day_code in shifts_day + ["日", MYONG_CODE]:
                if day_code in ALL_SHIFTS:
                    model.AddImplication(shifts[(s, d, MYONG_CODE)], shifts[(s, d + 1, day_code)].Not())

    # spacing night: d, d+2, d+4 <= 2
    for s in staff_indices:
        for d in range(num_days - 4):
            n1 = sum(shifts[(s, d, c)] for c in shifts_night if c in ALL_SHIFTS)
            n2 = sum(shifts[(s, d + 2, c)] for c in shifts_night if c in ALL_SHIFTS)
            n3 = sum(shifts[(s, d + 4, c)] for c in shifts_night if c in ALL_SHIFTS)
            model.Add(n1 + n2 + n3 <= 2)

    # 5 days window work <= 4
    for s in staff_indices:
        for d in range(num_days - 4):
            works = [1 - shifts[(s, d + k, OFF_CODE)] for k in range(5)]
            model.Add(sum(works) <= 4)

    # closed day: no night and no L1
    for d in closed_idx:
        for s in staff_indices:
            for c in shifts_night:
                model.Add(shifts[(s, d, c)] == 0)
            if "L1" in ALL_SHIFTS:
                model.Add(shifts[(s, d, "L1")] == 0)

    # HARD: user requests
    for s_idx, staff in enumerate(staff_data):
        name = staff["name"]
        if name not in requests:
            continue
        for day, req_code in requests[name].items():
            if 1 <= day <= num_days:
                d = day - 1
                if req_code in ALL_SHIFTS:
                    model.Add(shifts[(s_idx, d, req_code)] == 1)

    # ✅ HARD: night each code exactly 1 (non-closed)
    for d in days_indices:
        if d in closed_idx:
            continue
        for code in shifts_night:
            cnt = sum(shifts[(s, d, code)] for s in staff_indices)
            model.Add(cnt == 1)

    # ✅ HARD: L1 exactly 1 (non-closed)
    if "L1" in ALL_SHIFTS:
        for d in days_indices:
            if d in closed_idx:
                continue
            l1_cnt = sum(shifts[(s, d, "L1")] for s in staff_indices)
            model.Add(l1_cnt == 1)

    # Objective: prefer leaving unspecified day shifts as UNASSIGNED
    penalties = []
    requested_day_cells = set()
    for name, mp in requests.items():
        for day, code in mp.items():
            if code in shifts_day or code == "日":
                requested_day_cells.add((name, day))

    for s_idx, staff in enumerate(staff_data):
        name = staff["name"]
        for d in days_indices:
            day_num = d + 1
            if (name, day_num) in requested_day_cells:
                continue
            for code in shifts_day:
                if code in ALL_SHIFTS:
                    penalties.append(shifts[(s_idx, d, code)] * 2000)
            penalties.append(-50 * shifts[(s_idx, d, UNASSIGNED_CODE)])

    model.Minimize(sum(penalties))

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 10.0
    solver.parameters.num_search_workers = 8
    status = solver.Solve(model)
    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        return None, None

    day_headers = build_day_headers(year, month, num_days)

    schedule_data = []
    for s in staff_indices:
        row = {"Staff": staff_data[s]["name"]}
        work_days = 0
        off_days = 0
        for d in days_indices:
            val = None
            for code in ALL_SHIFTS:
                if solver.Value(shifts[(s, d, code)]):
                    val = code
                    break
            if val == OFF_CODE:
                off_days += 1
            else:
                work_days += 1
        row["公休数"] = off_days
        row["勤務日数(公以外)"] = work_days

        for d in days_indices:
            col = day_headers[d]
            val = "ERR"
            for code in ALL_SHIFTS:
                if solver.Value(shifts[(s, d, code)]):
                    val = code
                    break
            row[col] = "" if val == UNASSIGNED_CODE else val

        schedule_data.append(row)

    df_result = pd.DataFrame(schedule_data)
    df_summary = build_summary(df_result, staff_data, shifts_day, shifts_night, num_days, year, month, closed_idx)
    return df_result, df_summary

@st.cache_data(show_spinner=False)
def solve_stage2(num_days, year, month, prev_history, fixed_table, staff_data,
                shifts_day, shifts_night, closed_days, _version_stamp: str):
    """
    Stage2:
    - Stage1/수정본 고정값 하드
    - 빈칸 채워 완성
    - ✅ 야근(Q1,X1,R1) 각 1명 하드
    - ✅ L1 1명 하드
    """
    model = cp_model.CpModel()
    ALL_SHIFTS = shifts_day + shifts_night + SPECIAL_CODES
    staff_indices = range(len(staff_data))
    days_indices = range(num_days)
    closed_idx = set([d - 1 for d in closed_days if 1 <= d <= num_days])

    shifts = {}
    for s in staff_indices:
        for d in days_indices:
            for code in ALL_SHIFTS:
                shifts[(s, d, code)] = model.NewBoolVar(f"s2_s{s}_d{d}_{code}")

    for s in staff_indices:
        for d in days_indices:
            model.Add(sum(shifts[(s, d, c)] for c in ALL_SHIFTS) == 1)

    for s in staff_indices:
        allowed = parse_skills(staff_data[s].get("skills", ""))
        for d in days_indices:
            for code in ALL_SHIFTS:
                if code not in allowed:
                    model.Add(shifts[(s, d, code)] == 0)

    # prev month carry
    for s_idx, staff in enumerate(staff_data):
        name = staff["name"]
        h_d1 = norm_code(prev_history.get(name, {}).get("d-1", OFF_CODE))
        h_d2 = norm_code(prev_history.get(name, {}).get("d-2", OFF_CODE))
        h_d3 = norm_code(prev_history.get(name, {}).get("d-3", OFF_CODE))

        if h_d1 in shifts_night:
            model.Add(shifts[(s_idx, 0, MYONG_CODE)] == 1)

        if h_d1 == MYONG_CODE:
            for day_code in shifts_day + ["日", MYONG_CODE]:
                if day_code in ALL_SHIFTS:
                    model.Add(shifts[(s_idx, 0, day_code)] == 0)

        w_d3 = 1 if h_d3 != OFF_CODE else 0
        w_d2 = 1 if h_d2 != OFF_CODE else 0
        w_d1 = 1 if h_d1 != OFF_CODE else 0
        c0 = 1 - shifts[(s_idx, 0, OFF_CODE)] if 0 < num_days else 0
        c1 = 1 - shifts[(s_idx, 1, OFF_CODE)] if 1 < num_days else 0

        model.Add(w_d3 + w_d2 + w_d1 + c0 + c1 <= 4)
        if num_days >= 3:
            c2 = 1 - shifts[(s_idx, 2, OFF_CODE)]
            model.Add(w_d2 + w_d1 + c0 + c1 + c2 <= 4)
        if num_days >= 4:
            c2 = 1 - shifts[(s_idx, 2, OFF_CODE)]
            c3 = 1 - shifts[(s_idx, 3, OFF_CODE)]
            model.Add(w_d1 + c0 + c1 + c2 + c3 <= 4)

    # night -> next day is 明(-)
    for s in staff_indices:
        for d in range(num_days - 1):
            is_night = sum(shifts[(s, d, c)] for c in shifts_night if c in ALL_SHIFTS)
            model.Add(shifts[(s, d + 1, MYONG_CODE)] == is_night)

    # 明(-) -> next day cannot be day shift / 日 / 明
    for s in staff_indices:
        for d in range(num_days - 1):
            for day_code in shifts_day + ["日", MYONG_CODE]:
                if day_code in ALL_SHIFTS:
                    model.AddImplication(shifts[(s, d, MYONG_CODE)], shifts[(s, d + 1, day_code)].Not())

    # spacing night
    for s in staff_indices:
        for d in range(num_days - 4):
            n1 = sum(shifts[(s, d, c)] for c in shifts_night)
            n2 = sum(shifts[(s, d + 2, c)] for c in shifts_night)
            n3 = sum(shifts[(s, d + 4, c)] for c in shifts_night)
            model.Add(n1 + n2 + n3 <= 2)

    # 5 days window work<=4
    for s in staff_indices:
        for d in range(num_days - 4):
            works = [1 - shifts[(s, d + k, OFF_CODE)] for k in range(5)]
            model.Add(sum(works) <= 4)

    # closed day: no night and no L1
    for d in closed_idx:
        for s in staff_indices:
            for c in shifts_night:
                model.Add(shifts[(s, d, c)] == 0)
            if "L1" in ALL_SHIFTS:
                model.Add(shifts[(s, d, "L1")] == 0)

    # HARD: fixed_table (non-empty)
    day_headers = build_day_headers(year, month, num_days)
    name_to_idx = {s["name"]: i for i, s in enumerate(staff_data)}

    for _, r in fixed_table.iterrows():
        name = r["Staff"]
        if name not in name_to_idx:
            continue
        s_idx = name_to_idx[name]
        for d in days_indices:
            col = day_headers[d]
            v = norm_code(r.get(col, ""))
            if v == "":
                continue
            if v in ALL_SHIFTS:
                model.Add(shifts[(s_idx, d, v)] == 1)

    # ✅ HARD: night each code exactly 1 (non-closed)
    for d in days_indices:
        if d in closed_idx:
            continue
        for code in shifts_night:
            cnt = sum(shifts[(s, d, code)] for s in staff_indices)
            model.Add(cnt == 1)

    # ✅ HARD: L1 exactly 1 (non-closed)
    if "L1" in ALL_SHIFTS:
        for d in days_indices:
            if d in closed_idx:
                continue
            l1_cnt = sum(shifts[(s, d, "L1")] for s in staff_indices)
            model.Add(l1_cnt == 1)

    # Soft goals
    penalties = []

    # (E1+E2)+(G1+G1U) >= 2 (가능하면)
    e_codes = ["E1", "E2"]
    g_codes = ["G1", "G1U"]
    for d in days_indices:
        total_e = sum(shifts[(s, d, c)] for s in staff_indices for c in e_codes if c in ALL_SHIFTS)
        total_g = sum(shifts[(s, d, c)] for s in staff_indices for c in g_codes if c in ALL_SHIFTS)
        total_power = total_e + total_g
        is_short = model.NewBoolVar(f"s2_short_power_{d}")
        model.Add(total_power < 2).OnlyEnforceIf(is_short)
        model.Add(total_power >= 2).OnlyEnforceIf(is_short.Not())
        penalties.append(is_short * 50000)

    # Manager day >= 1 (가능하면)
    manager_indices = [i for i, s in enumerate(staff_data) if s["role"] == "Manager"]
    for d in days_indices:
        mgr_day = sum(shifts[(s, d, c)] for s in manager_indices for c in shifts_day if c in ALL_SHIFTS)
        is_zero = model.NewBoolVar(f"s2_mgr_zero_{d}")
        model.Add(mgr_day == 0).OnlyEnforceIf(is_zero)
        model.Add(mgr_day > 0).OnlyEnforceIf(is_zero.Not())
        penalties.append(is_zero * 50000)

    # OFF target (가능하면)
    for s in staff_indices:
        target_off = staff_data[s].get("target_off", 8)
        if pd.isna(target_off):
            target_off = 8
        target_off = int(target_off)

        actual_offs = model.NewIntVar(0, num_days, f"s2_off_{s}")
        model.Add(actual_offs == sum(shifts[(s, d, OFF_CODE)] for d in days_indices))

        diff = model.NewIntVar(0, num_days, f"s2_offdiff_{s}")
        model.AddAbsEquality(diff, actual_offs - target_off)
        penalties.append(diff * 100000)

    model.Minimize(sum(penalties))

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 10.0
    solver.parameters.num_search_workers = 8
    status = solver.Solve(model)
    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        return None, None

    schedule_data = []
    for s in staff_indices:
        row = {"Staff": staff_data[s]["name"]}
        off_days = 0
        work_days = 0

        for d in days_indices:
            val = "ERR"
            for code in ALL_SHIFTS:
                if solver.Value(shifts[(s, d, code)]):
                    val = code
                    break
            if val == OFF_CODE:
                off_days += 1
            else:
                work_days += 1

        row["公休数"] = off_days
        row["勤務日数(公以外)"] = work_days

        for d in days_indices:
            col = day_headers[d]
            val = "ERR"
            for code in ALL_SHIFTS:
                if solver.Value(shifts[(s, d, code)]):
                    val = code
                    break
            row[col] = val

        schedule_data.append(row)

    df_result = pd.DataFrame(schedule_data)
    df_summary = build_summary(df_result, staff_data, shifts_day, shifts_night, num_days, year, month, closed_idx)
    return df_result, df_summary

def build_summary(df_result, staff_data, shifts_day, shifts_night, num_days, year, month, closed_idx):
    day_headers = build_day_headers(year, month, num_days)
    daily_summary_list = []

    for d in range(num_days):
        col_name = day_headers[d]
        day_stats = {"日付": col_name}

        mgr_day = 0
        mgr_night = 0
        for s_idx, stf in enumerate(staff_data):
            v = df_result.iloc[s_idx][col_name]
            if stf["role"] == "Manager":
                if v in shifts_day: mgr_day += 1
                if v in shifts_night: mgr_night += 1

        day_stats["Manager(昼)"] = mgr_day
        day_stats["Manager(夜)"] = mgr_night

        all_codes = shifts_night + shifts_day + [OFF_CODE, MYONG_CODE, "日"]
        for code in all_codes:
            day_stats[code] = sum(1 for s_idx in range(len(staff_data)) if df_result.iloc[s_idx][col_name] == code)

        day_stats["休館"] = 1 if d in closed_idx else 0
        daily_summary_list.append(day_stats)

    return pd.DataFrame(daily_summary_list)

# =========================================================
# UI
# =========================================================
st.title("🏨 ホテルシフト自動作成 (2段階: Stage1→修正→Stage2 完成)")
st.caption(f"APP_VERSION: {APP_VERSION}")

with st.sidebar:
    st.header("⚙️ システム設定")

    if st.button("🔄 強制リセット（セッション＋キャッシュ）", key=versioned("force_reset")):
        reset_all_except_password()

    if st.button("ログアウト (Logout)", key=versioned("logout")):
        st.session_state["password_correct"] = False
        reset_all_except_password()

    st.header("📅 日付設定")
    col1, col2 = st.columns(2)
    year = col1.number_input("年", 2025, 2030, 2026, key=versioned("year"))
    month = col2.number_input("月", 1, 12, 1, key=versioned("month"))
    days_in_month = pd.Period(f"{year}-{month}").days_in_month
    st.info(f"計 {days_in_month}日")

with st.expander("⚙️ 勤務コード設定（新しい時間帯の追加・削除）"):
    st.caption("※ ✅ 夜勤は Q1/X1/R1 を毎日各1名固定。✅ L1も毎日1名固定。Dコードは廃止(空欄扱い)。")
    c1, c2 = st.columns(2)

    day_shifts_str = c1.text_area("日勤コード", ", ".join(st.session_state["shifts_day"]), key=versioned("day_shifts"))
    night_shifts_str = c2.text_area("夜勤コード（Q1,X1,R1の3つ推奨）", ", ".join(st.session_state["shifts_night"]), key=versioned("night_shifts"))

    st.session_state["shifts_day"] = [x.strip() for x in day_shifts_str.split(",") if x.strip()]
    st.session_state["shifts_night"] = [x.strip() for x in night_shifts_str.split(",") if x.strip()]
    remove_D_from_shift_lists()

    if len(st.session_state["shifts_night"]) != 3:
        st.warning("要求仕様: 夜勤は3名枠(Q1,X1,R1)固定。夜勤コードを3つにして。")
    if "L1" not in st.session_state["shifts_day"]:
        st.warning("L1が日勤コードにありません。L1は毎日1名必須です。")

    # Stage1 드롭다운 (희망휴일/희망근무/야근/L1/日 노출)
    DROPDOWN_STAGE1 = [""] + [OFF_CODE, "日"] + st.session_state["shifts_night"] + (["L1"] if "L1" in st.session_state["shifts_day"] else []) + st.session_state["shifts_day"]
    DROPDOWN_STAGE2 = [""] + [OFF_CODE, "日", MYONG_CODE] + st.session_state["shifts_night"] + st.session_state["shifts_day"]

with st.sidebar:
    st.divider()
    st.header("🏨 休館日")
    closed_days = st.multiselect(
        "休館日（この日は 夜勤 + L1 なし）",
        options=list(range(1, days_in_month + 1)),
        default=[],
        key=versioned("closed_days")
    )

with st.expander("👥 スタッフ管理（目標公休数＆可能勤務の編集）", expanded=True):
    df_staff = pd.DataFrame(INITIAL_STAFF_DB)
    edited_staff_df = st.data_editor(
        df_staff,
        num_rows="dynamic",
        column_config={
            "target_off": st.column_config.NumberColumn("目標公休数", min_value=0, max_value=31, step=1),
            "skills": st.column_config.TextColumn("可能勤務 (カンマ区切り)", width="large"),
            "role": st.column_config.SelectboxColumn("役職", options=["Manager", "Staff"]),
            "gender": st.column_config.SelectboxColumn("性別", options=["M", "F"]),
        },
        use_container_width=True,
        key=versioned("staff_editor"),
    )

current_names = edited_staff_df["name"].tolist() if "name" in edited_staff_df.columns else []

with st.expander("🔙 前月の最後3日間の勤務入力 (CSVアップロード対応)"):
    uploaded_prev = st.file_uploader("CSVファイルで一括アップロード (前月記録)", type=["csv"], key=versioned("prev_upload"))
    prev_cols = ["d-3", "d-2", "d-1"]

    if current_names:
        init_prev = pd.DataFrame(index=current_names, columns=prev_cols)
        if uploaded_prev is not None:
            try:
                df_upload_prev = pd.read_csv(uploaded_prev, index_col=0)
                for c in prev_cols:
                    if c in df_upload_prev.columns:
                        df_upload_prev[c] = df_upload_prev[c].map(norm_code)
                init_prev.update(df_upload_prev)
                st.success("CSVアップロード完了！")
            except Exception as e:
                st.error(f"CSV読み込みエラー: {e}")

        prev_column_config = {
            col: st.column_config.SelectboxColumn(col, width="small", options=DROPDOWN_STAGE2, required=False)
            for col in prev_cols
        }
        prev_editor = st.data_editor(init_prev, column_config=prev_column_config, num_rows="fixed", key=versioned("prev_editor"))
        st.download_button("📥 テンプレートをダウンロード (CSV)", init_prev.to_csv().encode("utf-8"), "prev_history_template.csv", key=versioned("prev_tpl"))
    else:
        st.warning("スタッフリストが空です。")
        prev_editor = pd.DataFrame()

st.divider()
st.subheader("Stage1：希望(公=希望休 / 希望勤務 / 夜勤(Q1,X1,R1) / L1 / 日)入力 → 自動でベース作成")

uploaded_req = st.file_uploader("CSV一括アップロード (Stage1 希望入力)", type=["csv"], key=versioned("stage1_req_upload"))
if current_names:
    init_data = pd.DataFrame(index=current_names, columns=[f"{i}日" for i in range(1, days_in_month + 1)])
    if uploaded_req is not None:
        try:
            df_upload_req = pd.read_csv(uploaded_req, index_col=0)
            for col in df_upload_req.columns:
                df_upload_req[col] = df_upload_req[col].map(norm_code)
            init_data.update(df_upload_req)
            st.success("CSVアップロード完了！")
        except Exception as e:
            st.error(f"CSV読み込みエラー: {e}")

    req_cfg = {
        col: st.column_config.SelectboxColumn(col, width="small", options=DROPDOWN_STAGE1, required=False)
        for col in init_data.columns
    }
    edited_stage1 = st.data_editor(init_data, column_config=req_cfg, num_rows="fixed", height=360, key=versioned("stage1_editor"))
    st.download_button("📥 テンプレートDL (CSV)", init_data.to_csv().encode("utf-8"), "stage1_request_template.csv", key=versioned("stage1_tpl"))
else:
    st.warning("スタッフリストが空です。")
    edited_stage1 = pd.DataFrame()

colA, colB = st.columns([1, 2])
with colA:
    run_stage1 = st.button("🚀 Stage1 自動作成", type="primary", key=versioned("run_stage1"))
with colB:
    st.info("※ ✅ 毎日 夜勤(Q1=1, X1=1, R1=1) ハード。✅ L1も毎日1名ハード。その他の主な日勤はStage2で完成。")

if run_stage1:
    if edited_staff_df.empty:
        st.error("スタッフデータがありません。")
        st.stop()
    if len(st.session_state["shifts_night"]) != 3:
        st.error("夜勤コードを3つにしてください（例: Q1, X1, R1）。")
        st.stop()
    if "L1" not in st.session_state["shifts_day"]:
        st.error("L1が日勤コードにありません。L1は毎日1名必須です。")
        st.stop()

    staff_data = edited_staff_df.to_dict("records")
    missing = validate_mandatory_coverage(staff_data, st.session_state["shifts_day"], st.session_state["shifts_night"])
    if missing:
        st.error(f"必須コードに対応できるスタッフが0人です: {', '.join(missing)}（スタッフのskillsを見直して）")
        st.stop()

    prev_history = {}
    if not prev_editor.empty:
        for staff_name in prev_editor.index:
            prev_history[staff_name] = {}
            for col in prev_cols:
                prev_history[staff_name][col] = norm_code(prev_editor.loc[staff_name, col]) or OFF_CODE

    requests = {}
    if not edited_stage1.empty:
        for staff_name in edited_stage1.index:
            requests[staff_name] = {}
            for day_col in edited_stage1.columns:
                v = norm_code(edited_stage1.loc[staff_name, day_col])
                if v == "":
                    continue
                day_num = int(day_col.replace("日", ""))
                requests[staff_name][day_num] = v

    st.write("### 🧾 Stage1 希望入力サマリー")
    st.dataframe(pd.DataFrame([summarize_requests(requests, st.session_state["shifts_day"], st.session_state["shifts_night"])]),
                 use_container_width=True)

    with st.spinner("Stage1（夜勤+L1+希望）を計算中..."):
        result_df1, summary_df1 = solve_stage1(
            days_in_month, year, month,
            prev_history, requests, staff_data,
            st.session_state["shifts_day"], st.session_state["shifts_night"],
            closed_days,
            APP_VERSION,  # ✅ 캐시 키에도 버전 반영
        )

    if result_df1 is None:
        st.error("❌ Stage1 실패: 조건 충돌(휴관/희망/야근 연속 규칙/스킬/인원 등)")
        st.stop()

    st.session_state["stage1_requests"] = requests
    st.session_state["stage1_staff_data"] = staff_data
    st.session_state["stage1_prev_history"] = prev_history
    st.session_state["stage1_result"] = result_df1

    st.success("✅ Stage1 완료! (주간 미정은 빈칸으로 남김)")
    st.write("### Stage1 결과")
    st.markdown(generate_colored_table_html(result_df1, requests), unsafe_allow_html=True)
    st.write("### Stage1 日別集計")
    st.dataframe(summary_df1, use_container_width=True, height=350)

st.divider()
st.subheader("修正（任意）→ Stage2：主な日勤も埋めて完成")

if "stage1_result" in st.session_state:
    st.caption("수정 안 하면 그대로 Stage2. 수정하면 그 값을 하드로 고정해서 Stage2가 나머지를 채움.")
    base_df = st.session_state["stage1_result"].copy()
    day_headers = build_day_headers(year, month, days_in_month)
    stage2_edit_cfg = {h: st.column_config.SelectboxColumn(h, width="small", options=DROPDOWN_STAGE2, required=False) for h in day_headers}

    edited_fixed = st.data_editor(
        base_df,
        num_rows="fixed",
        column_config=stage2_edit_cfg,
        use_container_width=True,
        height=420,
        key=versioned("stage2_fixed_editor"),
    )

    run_stage2 = st.button("✅ Stage2 完成させる", type="primary", key=versioned("run_stage2"))

    if run_stage2:
        staff_data = st.session_state["stage1_staff_data"]
        prev_history = st.session_state["stage1_prev_history"]
        requests = st.session_state["stage1_requests"]

        missing = validate_mandatory_coverage(staff_data, st.session_state["shifts_day"], st.session_state["shifts_night"])
        if missing:
            st.error(f"必須コードに対応できるスタッフが0人です: {', '.join(missing)}（スタッフのskillsを見直して）")
            st.stop()

        for h in day_headers:
            edited_fixed[h] = edited_fixed[h].map(norm_code)

        with st.spinner("Stage2（完成）を計算中..."):
            result_df2, summary_df2 = solve_stage2(
                days_in_month, year, month,
                prev_history, edited_fixed, staff_data,
                st.session_state["shifts_day"], st.session_state["shifts_night"],
                closed_days,
                APP_VERSION,  # ✅ 캐시 키 버전
            )

        if result_df2 is None:
            st.error("❌ Stage2 실패: 수정값이 규칙(야근→明, 휴관, 연속근무, 스킬)과 충돌했을 가능성 큼.")
            st.stop()

        st.success("✅ Stage2 완료! (최종 시프트)")
        st.write("### 📅 최종 시フト表")
        st.markdown(generate_colored_table_html(result_df2, requests), unsafe_allow_html=True)

        st.write("### 📊 日別集計")
        def highlight_zero(val):
            if isinstance(val, int) and val == 0:
                return "background-color: #ffcccc; color: red; font-weight: bold;"
            return ""
        st.dataframe(summary_df2.style.applymap(highlight_zero, subset=summary_df2.columns[1:]),
                     height=470, use_container_width=True)

        excel_data = create_styled_excel(result_df2, summary_df2, requests, year, month)
        st.download_button("📥 Excelダウンロード（色付き・集計・希望反映）",
                           excel_data, f"{year}_{month}_shift_styled.xlsx", key=versioned("dl_excel"))
else:
    st.info("Stage1을 먼저 실행해줘.")
